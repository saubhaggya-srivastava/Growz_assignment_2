"""Excel report generator with conditional formatting"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from .base import ReportGenerator
from ..models import ComparisonResult


class ExcelReportGenerator(ReportGenerator):
    """Generate reports in Excel format with conditional formatting"""
    
    def generate(self, comparison: ComparisonResult, output_path: str) -> None:
        """Generate Excel report with conditional formatting
        
        Args:
            comparison: ComparisonResult with discrepancies and summary
            output_path: Path where Excel file should be saved
        """
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write summary sheet
            if self.config.include_summary:
                self._write_summary_sheet(comparison, writer)
            
            # Write discrepancies sheet
            if self.config.include_matched_items:
                self._write_discrepancies_sheet(comparison, writer)
            
            # Write alerts sheet
            if self.config.include_alerts:
                self._write_alerts_sheet(comparison, writer)
        
        # Apply conditional formatting if requested
        if self.config.highlight_discrepancies:
            self._apply_conditional_formatting(output_path)
    
    def _write_summary_sheet(self, comparison: ComparisonResult, writer):
        """Write summary statistics sheet
        
        Args:
            comparison: ComparisonResult
            writer: Excel writer
        """
        summary_data = {
            'Metric': [
                'Total Quantity Ordered',
                'Total Quantity Invoiced',
                'Quantity Difference',
                'Total Value Ordered',
                'Total Value Invoiced',
                'Value Difference',
                'Items with Discrepancies',
                'Total Items Compared'
            ],
            'Value': [
                comparison.summary.total_quantity_ordered,
                comparison.summary.total_quantity_invoiced,
                comparison.summary.quantity_difference,
                f"${comparison.summary.total_value_ordered:.2f}",
                f"${comparison.summary.total_value_invoiced:.2f}",
                f"${comparison.summary.value_difference:.2f}",
                comparison.summary.items_with_discrepancies,
                len(comparison.discrepancies)
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _write_discrepancies_sheet(self, comparison: ComparisonResult, writer):
        """Write discrepancies sheet
        
        Args:
            comparison: ComparisonResult
            writer: Excel writer
        """
        rows = []
        
        for disc in comparison.discrepancies:
            # Calculate alert severity
            alert_severity = ''
            if disc.has_discrepancy and disc.po_item.total_value and disc.total_value_diff:
                percent_diff = abs(float(disc.total_value_diff / disc.po_item.total_value)) * 100
                if percent_diff >= 5.0:
                    alert_severity = 'HIGH'
                else:
                    alert_severity = 'LOW'
            
            row = {
                'PO SKU': disc.po_item.sku or '',
                'PO Description': disc.po_item.description,
                'PO Quantity': disc.po_item.quantity,
                'PO Unit Price': float(disc.po_item.unit_price) if disc.po_item.unit_price else None,
                'PO Total': float(disc.po_item.total_value) if disc.po_item.total_value else None,
                'PI SKU': disc.pi_item.sku or '',
                'PI Description': disc.pi_item.description,
                'PI Quantity': disc.pi_item.quantity,
                'PI Unit Price': float(disc.pi_item.unit_price) if disc.pi_item.unit_price else None,
                'PI Total': float(disc.pi_item.total_value) if disc.pi_item.total_value else None,
                'Qty Diff': disc.quantity_diff,
                'Price Diff': float(disc.unit_price_diff) if disc.unit_price_diff else None,
                'Total Diff': float(disc.total_value_diff) if disc.total_value_diff else None,
                'Has Discrepancy': 'Yes' if disc.has_discrepancy else 'No',
                'Qty Mismatch': 'Yes' if disc.has_quantity_mismatch else 'No',
                'Price Mismatch': 'Yes' if disc.has_price_mismatch else 'No',
                'Alert Severity': alert_severity
            }
            rows.append(row)
        
        df_disc = pd.DataFrame(rows)
        df_disc.to_excel(writer, sheet_name='Discrepancies', index=False)
    
    def _write_alerts_sheet(self, comparison: ComparisonResult, writer):
        """Write alerts sheet
        
        Args:
            comparison: ComparisonResult
            writer: Excel writer
        """
        alerts_data = {
            'Severity': [alert.severity for alert in comparison.alerts],
            'Message': [alert.message for alert in comparison.alerts],
            'Suggested Action': [alert.suggested_action for alert in comparison.alerts],
            'Related Item': [alert.related_item for alert in comparison.alerts]
        }
        
        df_alerts = pd.DataFrame(alerts_data)
        df_alerts.to_excel(writer, sheet_name='Alerts', index=False)
    
    def _apply_conditional_formatting(self, output_path: str):
        """Apply conditional formatting to highlight discrepancies
        
        Args:
            output_path: Path to Excel file
        """
        try:
            wb = load_workbook(output_path)
            
            # Format Discrepancies sheet
            if 'Discrepancies' in wb.sheetnames:
                ws = wb['Discrepancies']
                
                # Define fill colors
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                bold_font = Font(bold=True)
                
                # Find alert severity column
                header_row = [cell.value for cell in ws[1]]
                alert_severity_col = header_row.index('Alert Severity') + 1 if 'Alert Severity' in header_row else None
                
                # Apply formatting based on alert severity
                for row_idx in range(2, ws.max_row + 1):
                    severity = ''
                    
                    # Get alert severity
                    if alert_severity_col:
                        severity = ws.cell(row_idx, alert_severity_col).value or ''
                    
                    # Apply color based on severity
                    if severity == 'HIGH':
                        # Red for HIGH severity
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row_idx, col_idx).fill = red_fill
                            ws.cell(row_idx, col_idx).font = bold_font
                    elif severity == 'LOW':
                        # Yellow for LOW severity
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row_idx, col_idx).fill = yellow_fill
            
            # Format Alerts sheet
            if 'Alerts' in wb.sheetnames:
                ws = wb['Alerts']
                header_row = [cell.value for cell in ws[1]]
                severity_col = header_row.index('Severity') + 1 if 'Severity' in header_row else None
                
                if severity_col:
                    for row_idx in range(2, ws.max_row + 1):
                        severity = ws.cell(row_idx, severity_col).value
                        if severity in ['high', 'critical']:
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row_idx, col_idx).fill = red_fill
                                ws.cell(row_idx, col_idx).font = bold_font
            
            wb.save(output_path)
        
        except Exception as e:
            print(f"Warning: Could not apply conditional formatting: {e}")
